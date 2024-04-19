#--!Es necesario instalar la librería openpyxl!--
#--pip install openpyxl
import openpyxl
import random

# Solicitar al usuario los encabezados de las columnas
columnas = []
for i in range(1, 6):
  encabezado = input(f"Ingrese el encabezado de la columna {i}: ")
  columnas.append(encabezado)

# Crear un nuevo archivo de Excel
wb = openpyxl.Workbook()
ws = wb.active

# Escribir los encabezados en la fila 1
for i, encabezado in enumerate(columnas):
  ws.cell(row=1, column=i+1).value = encabezado

# Generar 100 filas de datos aleatorios
for fila in range(2, 102):
  for i, columna in enumerate(columnas):
    # Generar datos aleatorios según el tipo de columna
    if columna == "Número":
      dato = random.randint(1, 1000)
    elif columna == "Fecha":
      dato = random.date(2020, 1, 1, 2024, 12, 31)
    elif columna == "Cadena":
      dato = chr(random.randint(65, 90)) + chr(random.randint(97, 122)) + str(random.randint(100, 999))
    else:
      dato = "Dato no válido"

    ws.cell(row=fila, column=i+1).value = dato

# Guardar el archivo de Excel
nombre_archivo = input("Ingrese el nombre del archivo Excel: ")
wb.save(nombre_archivo + ".xlsx")

print(f"Datos aleatorios generados en el archivo: {nombre_archivo}.xlsx")